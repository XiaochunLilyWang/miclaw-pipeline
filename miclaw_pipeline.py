#!/usr/bin/env python3
"""
miclaw 舆情分析流水线

读取数据源 → 调用 Gemini 分析每条帖子（文字+图片+视频）→ 输出结构化 JSON

用法：
  python miclaw_pipeline.py inspect <excel>          # 查看列信息，输出 JSON
  python miclaw_pipeline.py run <source> [options]   # 运行分析

run 选项：
  --output <path>       结果输出路径（默认 pipeline_results.json）
  --source-label <s>    来源渠道标签（默认 小米社区）
  --resume              断点续跑
  --filter-col <name>   过滤列名（不传则处理全部行）
  --filter-val <vals>   过滤值，逗号分隔（配合 --filter-col 使用）
  --all-rows            忽略过滤，处理全部数据行
  --referer <url>       图片下载 Referer（默认 https://www.xiaomi.cn/）
"""

import sys
import os
import re
import json
import time
import base64
import argparse
import subprocess
import requests
from openai import OpenAI

# ── 读取 .env 文件（若存在）─────────────────────────────
def _load_dotenv():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        with open(env_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, _, v = line.partition("=")
                    os.environ.setdefault(k.strip(), v.strip())

_load_dotenv()

# ── 全局配置（优先读环境变量，其次 .env，再次默认值）────
GEMINI_API_KEY  = os.environ.get("MICLAW_GEMINI_API_KEY", "")
GEMINI_BASE_URL = os.environ.get("MICLAW_GEMINI_BASE_URL", "https://api.openai.com/v1/")
GEMINI_MODEL    = os.environ.get("MICLAW_GEMINI_MODEL",    "gpt-4o")

DEFAULT_REFERER = "https://www.xiaomi.cn/"
XIAOMI_API_BASE = "https://api.vip.miui.com/api/community/post/v2/detail"

SCENE_LABELS = [
    "IoT/智能家居控制",
    "系统级操作",
    "第三方App操作",
    "联网信息查询与整合",
    "记忆管理",
    "代码执行",
]

ANALYSIS_PROMPT = """
请分析这篇帖子，按如下 JSON 格式输出（不要有任何 markdown 包裹）：

{
  "sheet1": [
    {
      "场景类别": "IoT/智能家居控制 | 系统级操作 | 第三方App操作 | 联网信息查询与整合 | 记忆管理 | 代码执行",
      "用户指令": "用户对 miclaw 下达的具体指令，尽量还原原文",
      "执行情况": "AI 执行了什么操作/调用了什么工具/产出了什么结果",
      "任务是否成功": "成功 | 失败 | 部分成功",
      "用户评价": "用户对结果的评价（引用原话，无则留空）"
    }
  ],
  "sheet2": [
    {
      "评价对象": "被评价的具体方面（简洁标签）",
      "评价内容": "具体评价描述（保留用户原话关键内容）"
    }
  ]
}

判断规则：
- Sheet1：必须有明确的「用户指令 → AI 执行 → 结果」完整链路；一篇可多条
- Sheet2：主观评价、bug 报告、功能期待、版本更新、使用建议等；一篇可多条
- 同一篇可同时出现在两个 sheet
- 纯功能介绍/无实际操作 → 只有 sheet2
- 没有 sheet1 内容时 sheet1 返回 []，没有 sheet2 内容时 sheet2 返回 []
"""

ABILITY_TAGGING_PROMPT = """
为以下 miclaw 使用 case 打能力标签（每个 case 可打多个，同一标签只出现一次）。

## 打标规则

| 触发条件 | 标签 |
|---------|------|
| 读取/查找/定位数据（短信、联系人、文件、图片内容等） | 信息查找/筛选 |
| 图片/音频/屏幕画面特征提取 | 多模态解析 |
| 智能家居传感器数据读取（温度、湿度、水量等） | 环境感知 |
| 对内容进行总结、归纳、提炼要点，或提取验证码/取件码等关键要素 | 信息总结与归纳 |
| 涉及数量/金额/频次计算 | 数据统计与计算 |
| 重新排版/格式转换 | 内容重组与排版 |
| 触发系统底层功能（发短信/打电话/设闹钟/截图/录屏等） | 系统操作执行 |
| 调用第三方 App / 执行代码 / 配置 MCP / 创建 skill | 工具调用和创造 |
| 同时调度多个外部应用（二元交互不算） | 跨应用协同调度 |
| 控制单个智能家居设备（开关/调节/启动等） | 智能设备操作执行 |
| 同时控制或联动多个智能家居设备 | 多设备协同调度 |
| 将模糊指令拆解为多个操作步骤 | 多步骤任务规划 |
| 定时/事件触发自动执行 | 自动化 |
| 记住用户习惯/偏好/人设 | 记忆 |
| 多方案尝试 / 用户不满后快速调整 / 主动向用户确认再执行 | 自我反思和任务闭环 |

## 重要注意事项

- 总结/提取短信 → 必须同时打：信息查找/筛选 + 信息总结与归纳 + 系统操作执行
- 仅涉及 miclaw 与另一个应用的二元交互 → 不打「跨应用协同调度」
- 可用标签（严格使用以下名称，不要自创）：
  信息查找/筛选、多模态解析、环境感知、信息总结与归纳、数据统计与计算、内容重组与排版、
  多步骤任务规划、系统操作执行、跨应用协同调度、工具调用和创造、智能设备操作执行、多设备协同调度、
  自动化、记忆、自我反思和任务闭环

## 待打标 Cases

{cases_json}

## 输出格式

返回 JSON 数组，顺序与输入对应，每个元素只含 case_index 和 能力标签：
[{"case_index": 0, "能力标签": ["标签A", "标签B"]}, ...]
不要有任何 markdown 包裹。
"""

# ── 工具函数 ──────────────────────────────────────────

def log(msg): print(msg, flush=True)

def sanitize_text(s):
    """替换影响飞书 MCP JSON 解析的 Unicode 弯引号为直引号"""
    if not isinstance(s, str):
        return s
    return s.replace('\u201c', '"').replace('\u201d', '"')


def build_export_data(results):
    """
    从 pipeline_results.json 构建适合写入飞书的二维数组。
    返回 (s1_rows, s2_rows)，含表头，文本已净化弯引号。
    """
    S1_HEADER = ["来源渠道", "作者", "原文链接", "场景类别", "用户指令", "执行情况", "任务是否成功", "用户评价", "能力标签"]
    S2_HEADER = ["来源渠道", "作者", "原文链接", "评价对象", "评价内容"]

    def c(v):
        return sanitize_text(str(v) if v is not None else "")

    s1_rows = [S1_HEADER]
    s2_rows = [S2_HEADER]

    for item in results:
        label  = c(item.get("source_label", ""))
        author = c(item.get("作者", ""))
        url    = c(item.get("url", ""))

        for entry in item.get("sheet1", []):
            ability_tags = entry.get("能力标签", [])
            ability_str = c("、".join(ability_tags) if isinstance(ability_tags, list) else str(ability_tags))
            s1_rows.append([
                label, author, url,
                c(entry.get("场景类别", "")),
                c(entry.get("用户指令", "")),
                c(entry.get("执行情况", "")),
                c(entry.get("任务是否成功", "")),
                c(entry.get("用户评价", "")),
                ability_str,
            ])

        for entry in item.get("sheet2", []):
            s2_rows.append([
                label, author, url,
                c(entry.get("评价对象", "")),
                c(entry.get("评价内容", "")),
            ])

    return s1_rows, s2_rows


def tag_abilities_for_post(client, cases):
    """
    为一篇帖子的所有 sheet1 cases 打能力标签。
    cases: [{"case_index": i, "用户指令": "...", "执行情况": "..."}, ...]
    返回: {case_index: [标签列表], ...}
    """
    if not cases:
        return {}
    cases_json = json.dumps(cases, ensure_ascii=False, indent=2)
    prompt = ABILITY_TAGGING_PROMPT.replace("{cases_json}", cases_json)
    resp = client.chat.completions.create(
        model=GEMINI_MODEL,
        messages=[{"role": "user", "content": prompt}],
        max_tokens=1000,
    )
    raw = resp.choices[0].message.content.strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
    results = json.loads(raw)
    return {item["case_index"]: item["能力标签"] for item in results}


def load_progress(path):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"done": [], "failed": [], "results": []}

def save_progress(data, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── 数据读取 ──────────────────────────────────────────

def inspect_excel(path):
    """
    检查 Excel 文件，返回列信息 JSON：
    - 总行数
    - 每列：列名、非空值数、唯一值数、前 10 个高频值（带计数）
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要 openpyxl：pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]

    # 收集每列的值
    from collections import Counter
    col_values = {h: [] for h in headers}
    total_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        for i, v in enumerate(row):
            if i < len(headers):
                col_values[headers[i]].append(str(v).strip() if v is not None else "")

    wb.close()

    columns = []
    for h in headers:
        vals = col_values[h]
        non_empty = [v for v in vals if v not in ("", "None")]
        counter = Counter(non_empty)
        top = counter.most_common(10)
        columns.append({
            "name": h,
            "non_empty": len(non_empty),
            "unique": len(counter),
            "top_values": [{"value": v, "count": c} for v, c in top],
        })

    return {"file": path, "total_rows": total_rows, "columns": columns}


def load_from_excel(path, filter_col=None, filter_vals=None):
    """
    从 Excel 读取需要处理的行，返回 dict 列表。
    filter_col=None 或 filter_vals=None 时返回全部数据行。
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("需要 openpyxl：pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else f"col_{i}"
               for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)}
        if filter_col and filter_vals:
            if row_dict.get(filter_col, "") not in filter_vals:
                continue
        rows.append(row_dict)
    return rows


def is_xhs_json(records):
    """检测是否为小红书 MediaCrawler JSON 格式"""
    return bool(records) and isinstance(records[0], dict) and "note_id" in records[0]


def normalize_xhs_row(row):
    """将小红书 JSON 字段映射为 pipeline 内部字段名（保留原始字段）"""
    normalized = dict(row)
    normalized.update({
        "序号":     row.get("note_id", ""),
        "作者":     row.get("nickname", ""),
        "标题":     row.get("title", ""),
        "解析正文": row.get("desc", ""),
        "图片URL":  row.get("image_list", ""),
        "视频URL":  row.get("video_url", ""),
        "原文链接": row.get("note_url", ""),
        "_source":  "xhs",
    })
    return normalized


def load_from_json(path, filter_col=None, filter_vals=None):
    with open(path, "r", encoding="utf-8") as f:
        records = json.load(f)

    if is_xhs_json(records):
        log(f"  检测到小红书 JSON 格式，共 {len(records)} 条")
        records = [normalize_xhs_row(r) for r in records]
        # 默认只处理 是否分析=1 的记录（除非调用方已指定过滤条件）
        if filter_col is None:
            filter_col = "是否分析"
            filter_vals = ("1",)

    if filter_col and filter_vals:
        before = len(records)
        records = [r for r in records if str(r.get(filter_col, "")) in [str(v) for v in filter_vals]]
        log(f"  过滤 {filter_col}∈{filter_vals}：{before} → {len(records)} 条")

    return records


def fetch_single_post(post_id, source_label="小米社区"):
    """通过小米社区 API 获取单条帖子数据"""
    headers = {
        "User-Agent": "Mozilla/5.0 (Linux; Android 14; 23116PN5BC) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36",
        "Referer": "https://web.vip.miui.com/",
        "Origin": "https://web.vip.miui.com",
    }
    params = {"postId": post_id, "pathname": "/mio/detail", "version": "dev.20051", "fromBoardId": "undefined"}
    resp = requests.get(XIAOMI_API_BASE, params=params, headers=headers, timeout=15)
    resp.raise_for_status()
    data = resp.json()
    # 尝试从响应中提取结构
    post = data.get("data", {}).get("post", data.get("data", {}))
    return {
        "序号": post_id,
        "作者": post.get("authorInfo", {}).get("name", "") if isinstance(post, dict) else "",
        "标题": post.get("title", "") if isinstance(post, dict) else "",
        "解析正文": post.get("content", "") if isinstance(post, dict) else "",
        "原文链接": f"https://www.xiaomi.cn/post/{post_id}",
        "图片URL": "",
        "视频URL": "",
        "_raw": data,
    }


def resolve_source(source, filter_col, filter_vals, source_label):
    """解析 source 参数，返回 (rows, source_label)"""
    # URL 或纯 postId
    if re.search(r'(?:postId[=/]|/post/)(\d+)', source):
        m = re.search(r'(?:postId[=/]|/post/)(\d+)', source)
        return [fetch_single_post(m.group(1), source_label)], source_label
    if re.fullmatch(r'\d+', source):
        return [fetch_single_post(source, source_label)], source_label

    if source.endswith(".json"):
        return load_from_json(source, filter_col, filter_vals), source_label

    if source.endswith((".xlsx", ".xls")):
        return load_from_excel(source, filter_col, filter_vals), source_label

    raise ValueError(f"无法识别数据源：{source}")


# ── 图片/视频下载 ──────────────────────────────────────

def download_images(urls_str, referer=DEFAULT_REFERER, max_count=5):
    if not urls_str or urls_str.strip() in ("", "None"):
        return []
    # 支持空格分隔（小米社区）和逗号分隔（小红书图文笔记）两种格式
    raw = urls_str.replace(",", " ")
    urls = [u.strip() for u in raw.split() if u.strip().startswith("http")]
    images = []
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36", "Referer": referer}
    for url in urls[:max_count]:
        try:
            resp = requests.get(url, headers=headers, timeout=20)
            if resp.status_code == 200 and len(resp.content) > 500:
                ct = resp.headers.get("content-type", "")
                ext = "webp" if ("webp" in ct or "webp" in url) else ("png" if "png" in ct else "jpeg")
                images.append((ext, base64.b64encode(resp.content).decode()))
        except Exception as e:
            log(f"    图片下载失败: {url[:60]}: {e}")
    return images


def download_video_for_gemini(video_url, post_id, work_dir, max_mb=20, max_parts=3):
    """
    下载视频，若超过 max_mb 则用 FFmpeg 分段压缩。
    最多返回 max_parts 段（避免 base64 请求体过大触发 413）。
    返回本地文件路径列表（供 Gemini 分析）。
    """
    raw_path = os.path.join(work_dir, f"video_{post_id}_raw.mp4")
    h = {"User-Agent": "Mozilla/5.0 (Linux; Android 14) AppleWebKit/537.36", "Referer": "https://web.vip.miui.com/"}
    log(f"    下载视频: {video_url[:80]}...")
    with requests.get(video_url, headers=h, stream=True, timeout=120) as r:
        r.raise_for_status()
        with open(raw_path, "wb") as f:
            for chunk in r.iter_content(65536):
                f.write(chunk)
    size_mb = os.path.getsize(raw_path) / 1024 / 1024
    log(f"    视频下载完成: {size_mb:.1f} MB")

    if size_mb <= max_mb:
        return [raw_path]

    # 分段压缩
    probe = subprocess.run(
        ["ffprobe", "-v", "quiet", "-print_format", "json", "-show_format", raw_path],
        capture_output=True, text=True
    )
    duration = float(json.loads(probe.stdout)["format"]["duration"])
    n_parts = max(2, int(size_mb / (max_mb * 0.8)) + 1)
    # 只压缩前 max_parts 段，避免请求体过大（413）
    n_compress = min(n_parts, max_parts)
    seg_secs = duration / n_parts
    base = os.path.join(work_dir, f"video_{post_id}")
    parts = []
    for i in range(n_compress):
        out = f"{base}_part{i+1}.mp4"
        subprocess.run([
            "ffmpeg", "-y", "-ss", str(i * seg_secs), "-t", str(seg_secs),
            "-i", raw_path, "-vf", "scale=640:-2",
            "-c:v", "libx264", "-b:v", "400k", "-c:a", "aac", "-b:a", "64k", out
        ], capture_output=True)
        parts.append(out)
        log(f"    压缩第 {i+1}/{n_compress} 段（共{n_parts}段，截取前{n_compress}）: {os.path.getsize(out)/1024/1024:.1f} MB")
    if n_compress < n_parts:
        log(f"    注意：视频过长，仅分析前 {n_compress} 段（约前 {n_compress*seg_secs:.0f}s）")
    os.remove(raw_path)
    return parts


def get_xiaomi_video_url(post_id):
    """通过小米社区 API 获取帖子的真实视频 CDN 链接"""
    headers = {
        "User-Agent": "Mozilla/5.0 (Linux; Android 14; 23116PN5BC) AppleWebKit/537.36",
        "Referer": "https://web.vip.miui.com/",
    }
    params = {"postId": post_id, "pathname": "/mio/detail", "version": "dev.20051"}
    resp = requests.get(XIAOMI_API_BASE, params=params, headers=headers, timeout=15)
    resp.raise_for_status()
    body = resp.text
    mp4 = next(iter(re.findall(r'https?://[^\s"\']+\.mp4[^\s"\']*', body)), "")
    m3u8 = next(iter(re.findall(r'https?://[^\s"\']+\.m3u8[^\s"\']*', body)), "")
    return mp4 or m3u8


# ── Gemini 分析 ───────────────────────────────────────

def analyze_post(client, row, images, video_files=None, source_label="小米社区"):
    """调用 Gemini 分析单条帖子，返回 {sheet1: [...], sheet2: [...]}"""

    # 拼接文本
    text = f"""帖子信息：
来源：{source_label}
作者：{row.get('作者', '')}
标题：{row.get('标题', '')}
正文：{row.get('解析正文', '') or row.get('摘要', '')}"""

    # 小红书额外字段
    if row.get("_source") == "xhs":
        xhs_extra = []
        if row.get("type"):
            xhs_extra.append(f"内容类型：{row['type']}")
        if row.get("liked_count"):
            xhs_extra.append(f"点赞：{row['liked_count']}  收藏：{row.get('collected_count','')}  评论：{row.get('comment_count','')}")
        if xhs_extra:
            text += "\n" + "\n".join(xhs_extra)
    else:
        text += f"""
类型：{row.get('类型', '')}
情感：{row.get('情感', '')}
评价模块：{row.get('评价模块', '')}"""

    # 附上已有预分析字段（如有）
    pre = {k: row[k] for k in ("用户指令", "执行情况", "场景类别", "用户评价", "执行成功")
           if row.get(k) and row.get(k) not in ("None", "信息不足，无法判断", "")}
    if pre:
        text += f"\n\n已有分析（供参考）：{json.dumps(pre, ensure_ascii=False)}"

    text += ANALYSIS_PROMPT

    # 构建多模态 content
    content = []
    for ext, b64 in images:
        content.append({"type": "image_url", "image_url": {"url": f"data:image/{ext};base64,{b64}"}})

    # 若有视频，逐段读入 base64（用 video_url 类型，image_url 不支持 mp4）
    for vpath in (video_files or []):
        try:
            with open(vpath, "rb") as vf:
                vb64 = base64.b64encode(vf.read()).decode()
            content.append({"type": "video_url", "video_url": {"url": f"data:video/mp4;base64,{vb64}"}})
        except Exception as e:
            log(f"    视频编码失败: {e}")

    content.append({"type": "text", "text": text})

    resp = client.chat.completions.create(
        model=GEMINI_MODEL,
        messages=[{"role": "user", "content": content}],
        max_tokens=2000,
    )
    raw = resp.choices[0].message.content.strip()

    # 去除 markdown 包裹
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)

    return json.loads(raw)


# ── 主流程 ────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="miclaw 帖子分析流水线")
    subparsers = parser.add_subparsers(dest="cmd")

    # ── inspect 子命令 ──────────────────────────────
    p_inspect = subparsers.add_parser("inspect", help="查看 Excel 列信息")
    p_inspect.add_argument("excel", help="Excel 文件路径")

    # ── prepare-export 子命令 ────────────────────────
    p_export = subparsers.add_parser("prepare-export", help="从分析结果生成飞书导入数据（净化弯引号）")
    p_export.add_argument("input", nargs="?", default="pipeline_results.json",
                          help="分析结果 JSON（默认 pipeline_results.json）")
    p_export.add_argument("--s1-output", default="s1_data.json")
    p_export.add_argument("--s2-output", default="s2_data.json")

    # ── tag-abilities 子命令 ─────────────────────────
    p_tag = subparsers.add_parser("tag-abilities", help="为 sheet1 cases 打能力标签")
    p_tag.add_argument("input", nargs="?", default="pipeline_results.json",
                       help="分析结果 JSON（默认 pipeline_results.json）")
    p_tag.add_argument("--output", default=None, help="输出路径（默认覆盖原文件）")

    # ── run 子命令 ──────────────────────────────────
    p_run = subparsers.add_parser("run", help="运行分析流水线")
    p_run.add_argument("source", help="数据源：Excel/JSON 路径、帖子 URL 或 postId")
    p_run.add_argument("--output", default="pipeline_results.json")
    p_run.add_argument("--source-label", default="小米社区")
    p_run.add_argument("--resume", action="store_true")
    p_run.add_argument("--filter-col", default=None, help="过滤列名（不传则处理全部行）")
    p_run.add_argument("--filter-val", default=None, help="过滤值，逗号分隔")
    p_run.add_argument("--all-rows", action="store_true", help="忽略过滤，处理全部数据行")
    p_run.add_argument("--referer", default=DEFAULT_REFERER)

    args = parser.parse_args()

    # ── inspect ─────────────────────────────────────
    if args.cmd == "inspect":
        if args.excel.endswith(".json"):
            with open(args.excel, "r", encoding="utf-8") as f:
                records = json.load(f)
            from collections import Counter
            total = len(records)
            columns = []
            if records:
                keys = list(records[0].keys())
                for k in keys:
                    vals = [str(r.get(k, "")).strip() for r in records]
                    non_empty = [v for v in vals if v not in ("", "None")]
                    counter = Counter(non_empty)
                    columns.append({
                        "name": k,
                        "non_empty": len(non_empty),
                        "unique": len(counter),
                        "top_values": [{"value": v, "count": c} for v, c in counter.most_common(5)],
                    })
            result = {"file": args.excel, "total_rows": total, "columns": columns}
        else:
            result = inspect_excel(args.excel)
        out_path = "inspect_result.json"  # relative to CWD; read via file tool to avoid GBK
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print("INSPECT_OUTPUT: inspect_result.json")
        return

    # ── prepare-export ──────────────────────────────
    if args.cmd == "prepare-export":
        with open(args.input, "r", encoding="utf-8") as f:
            results = json.load(f)
        s1_rows, s2_rows = build_export_data(results)
        with open(args.s1_output, "w", encoding="utf-8") as f:
            json.dump(s1_rows, f, ensure_ascii=False)
        with open(args.s2_output, "w", encoding="utf-8") as f:
            json.dump(s2_rows, f, ensure_ascii=False)
        log(f"导出完成：")
        log(f"  Sheet1（任务执行）：{len(s1_rows)-1} 条 → {args.s1_output}")
        log(f"  Sheet2（用户评价）：{len(s2_rows)-1} 条 → {args.s2_output}")
        return

    # ── tag-abilities ────────────────────────────────
    if args.cmd == "tag-abilities":
        if not GEMINI_API_KEY:
            print("错误：未找到 MICLAW_GEMINI_API_KEY")
            sys.exit(1)
        with open(args.input, "r", encoding="utf-8") as f:
            results = json.load(f)
        output_path = args.output or args.input
        client = OpenAI(api_key=GEMINI_API_KEY, base_url=GEMINI_BASE_URL)
        total_cases = sum(len(item.get("sheet1", [])) for item in results)
        tagged_cases = 0
        for i, item in enumerate(results):
            s1 = item.get("sheet1", [])
            if not s1:
                continue
            log(f"[{i+1}/{len(results)}] 序号={item.get('序号', i)}  作者={item.get('作者', '')}  ({len(s1)} cases)")
            cases_input = [
                {"case_index": j, "用户指令": c.get("用户指令", ""), "执行情况": c.get("执行情况", "")}
                for j, c in enumerate(s1)
            ]
            try:
                tag_map = tag_abilities_for_post(client, cases_input)
                for j, c in enumerate(s1):
                    c["能力标签"] = tag_map.get(j, [])
                tagged_cases += len(s1)
                log(f"  → {[c['能力标签'] for c in s1]}")
            except Exception as e:
                log(f"  打标失败: {e}")
            time.sleep(1)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        log(f"\n完成：{tagged_cases}/{total_cases} 条 cases 已打标 → {output_path}")
        return

    # ── run ─────────────────────────────────────────
    if args.cmd != "run":
        parser.print_help()
        return

    # 检查必要配置
    if not GEMINI_API_KEY:
        print("错误：未找到 MICLAW_GEMINI_API_KEY")
        print("请在项目目录创建 .env 文件，参考 .env.example")
        sys.exit(1)

    filter_col = None if args.all_rows else args.filter_col
    filter_vals = None
    if filter_col and args.filter_val:
        filter_vals = tuple(v.strip() for v in args.filter_val.split(","))
    work_dir = os.path.dirname(os.path.abspath(args.output))
    progress_file = args.output.replace(".json", "_progress.json")

    # 1. 加载数据
    log(f"加载数据源：{args.source}")
    rows, source_label = resolve_source(args.source, filter_col, filter_vals, args.source_label)
    log(f"共 {len(rows)} 条待处理")

    # 2. 断点续跑
    progress = load_progress(progress_file) if args.resume else {"done": [], "failed": [], "results": []}
    done_ids = set(progress["done"])
    results = progress["results"]

    # 3. 初始化 Gemini client
    client = OpenAI(api_key=GEMINI_API_KEY, base_url=GEMINI_BASE_URL)

    # 4. 逐条处理
    for i, row in enumerate(rows):
        seq_id = str(row.get("序号", i))
        url = row.get("原文链接", row.get("url", ""))
        author = row.get("作者", "")

        if seq_id in done_ids:
            log(f"[{i+1}/{len(rows)}] 序号{seq_id} 已完成，跳过")
            continue

        log(f"\n[{i+1}/{len(rows)}] 序号={seq_id}  作者={author}")

        # 下载图片（小红书图片用 xhs CDN referer）
        referer = "https://www.xiaohongshu.com/" if row.get("_source") == "xhs" else args.referer
        images = download_images(row.get("图片URL", ""), referer=referer)
        if images:
            log(f"  图片：{len(images)} 张")

        # 下载视频（若有）
        video_files = []
        vid_url = row.get("视频URL", "").strip()
        post_id = seq_id

        # 尝试通过 API 获取视频（小米社区）
        if not vid_url and "xiaomi.cn/post/" in url:
            m = re.search(r'/post/(\d+)', url)
            if m:
                try:
                    vid_url = get_xiaomi_video_url(m.group(1))
                    if vid_url:
                        log(f"  发现视频（API获取）: {vid_url[:60]}")
                except Exception as e:
                    log(f"  获取视频URL失败: {e}")

        if vid_url:
            try:
                video_files = download_video_for_gemini(vid_url, post_id, work_dir)
            except Exception as e:
                log(f"  视频下载失败: {e}")

        # 分析（视频不支持时自动降级为纯文字+图片）
        try:
            try:
                result = analyze_post(client, row, images, video_files, source_label)
            except Exception as e:
                if video_files and ("video" in str(e).lower() or "image type" in str(e).lower() or "unsupported" in str(e).lower()):
                    log(f"  视频格式不支持，降级为仅文字+图片重试...")
                    result = analyze_post(client, row, images, [], source_label)
                else:
                    raise
            s1 = result.get("sheet1", [])
            s2 = result.get("sheet2", [])
            log(f"  → Sheet1: {len(s1)}条, Sheet2: {len(s2)}条")

            results.append({
                "序号": seq_id,
                "作者": author,
                "url": url,
                "source_label": source_label,
                "sheet1": s1,
                "sheet2": s2,
            })
            done_ids.add(seq_id)
            progress["done"] = list(done_ids)
            progress["results"] = results
            save_progress(progress, progress_file)

        except Exception as e:
            log(f"  分析失败: {e}")
            progress["failed"] = progress.get("failed", []) + [seq_id]
            save_progress(progress, progress_file)

        # 清理视频分段
        for vf in video_files:
            try:
                os.remove(vf)
            except:
                pass

        time.sleep(1.5)

    # 5. 输出最终结果
    save_progress({"done": list(done_ids), "failed": progress.get("failed", []), "results": results}, progress_file)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    log(f"\n=== 完成 ===")
    log(f"成功：{len(done_ids)} 条，失败：{len(progress.get('failed', []))} 条")
    log(f"结果已保存至：{args.output}")


if __name__ == "__main__":
    main()
